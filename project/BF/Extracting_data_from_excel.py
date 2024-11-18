import json
import openpyxl
from sentence_transformers import SentenceTransformer
import faiss
import numpy as np
import os


# Configuration
file_path = "C:\\Users\\I1817\\Downloads\\book21.xlsx" # Update this path
index_save_path = "D:/unsloth/Indium/project/BF/faiss_index.index"  # Update this path
metadata_save_path = "D:/unsloth/Indium/project/BF/meradata.json"  # Update this path
embedding_dimension = 384  # Dimensions of the sentence embeddings
model_name = "sentence-transformers/all-MiniLM-L6-v2"

# Load the Excel file
try:
    workbook = openpyxl.load_workbook(file_path)
except Exception as e:
    print(f"Error loading the Excel file: {e}")
    exit()

# Initialize Sentence Transformer model
model = SentenceTransformer(model_name)

# Create a FAISS index
index = faiss.IndexFlatL2(embedding_dimension)

# Store metadata
metadata = []

# Extract data from Excel (starting with Sheet2)
sheet_name = "Sheet2"  # Update this if your sheet is named differently
if sheet_name not in workbook.sheetnames:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
    exit()

sheet2 = workbook[sheet_name]
class_data = {}

# Fix: Use max_col=4 as we expect 4 columns
for row in sheet2.iter_rows(min_row=2, max_col=4, values_only=True):
    folder, class_name, class_description, method_tag = row

    if class_name and class_description and method_tag:
        # Store class metadata
        class_data[method_tag] = {
            "folder": folder,
            "class_name": class_name,
            "class_description": class_description,
            "methods": {}
        }

# Check if class_data is populated
print(f"Classes extracted: {len(class_data)}")

# Process each method sheet (M001, M002...)
for method_tag in class_data.keys():
    if method_tag not in workbook.sheetnames:
        print(f"Method sheet '{method_tag}' not found in the workbook.")
        continue

    method_sheet = workbook[method_tag]

    for row in method_sheet.iter_rows(min_row=2, values_only=True):
        # Unpack only the first four elements from the row, ignoring extra values
        serial_number, method_name, code_implementation, method_description = row[:4]

        if method_name:
            # Store method metadata
            class_data[method_tag]["methods"][method_name] = {
                "serial_number": serial_number,
                "code_implementation": code_implementation,
                "method_description": method_description
            }

            # Concatenate method description and implementation
            combined_text = f"{method_description} {code_implementation}"

            # Create embedding for combined text (description + implementation)
            embedding = model.encode(combined_text)

            # Add to FAISS index
            index.add(np.array([embedding]))

            # Store the corresponding metadata for retrieval
            metadata.append({
                "folder": class_data[method_tag]["folder"],
                "class_name": class_data[method_tag]["class_name"],
                "method_name": method_name,
                "method_description": method_description,
                "code_implementation": code_implementation  # Include implementation in metadata
            })

            # Debug print for each method added
            print(f"Added method: {method_name} with description and implementation")

# Save FAISS index and metadata for later retrieval
try:
    faiss.write_index(index, index_save_path)
    with open(metadata_save_path, "w") as f:
        json.dump(metadata, f, indent=4)
    print("Data extraction and vectorization completed. Index saved.")
except Exception as e:
    print(f"Error saving index or metadata: {e}")

# Print final count of methods in metadata
print(f"Total methods in metadata: {len(metadata)}")
